# services/export_service.py
import openpyxl
import tempfile
import os
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from models.export import ExportJob
from models.assignment import Assignment
from models.user import User
from models.submission import Submission
from models.class_ import ClassMembership
from services.s3_service import s3
from datetime import datetime

def generate_export(export_job_id: str, db: Session):
    job = db.query(ExportJob).filter_by(id=export_job_id).first()
    if not job:
        return

    try:
        assignment = db.query(Assignment).filter_by(id=job.assignment_id).first()

        # fetch all active students + their current submissions
        students = db.query(User, Submission).join(
            ClassMembership,
            (ClassMembership.user_id == User.id) &
            (ClassMembership.class_id == assignment.class_id) &
            (ClassMembership.member_role == 'STUDENT') &
            (ClassMembership.status == 'ACTIVE')
        ).outerjoin(
            Submission,
            (Submission.student_id == User.id) &
            (Submission.assignment_id == job.assignment_id) &
            (Submission.is_current == True)
        ).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tracker"

        # Header row
        headers = [
            "Student Name", "Registration ID", "Status",
            "Submitted At", "Is Late", "Version", "Submission Type"
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E4057")
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        status_colors = {
            "SUBMITTED": "C8E6C9",
            "LATE": "FFE0B2",
            "MISSED": "FFCDD2",
            "PENDING": "F5F5F5"
        }
        for row_idx, (student, sub) in enumerate(students, start=2):
            if sub:
                status = "LATE" if sub.is_late else "SUBMITTED"
                submitted_at = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else ""
                is_late = "Yes" if sub.is_late else "No"
                version = sub.version
                sub_type = sub.submission_type
            else:
                status = "MISSED" if assignment.status == "CLOSED" else "PENDING"
                submitted_at = ""
                is_late = ""
                version = ""
                sub_type = ""

            row_data = [
                student.full_name, student.registration_id,
                status, submitted_at, is_late, version, sub_type
            ]
            fill = PatternFill("solid", fgColor=status_colors.get(status, "FFFFFF"))
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill = fill

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        from models.analytics import AssignmentAnalytics
        aa = db.query(AssignmentAnalytics).filter_by(assignment_id=job.assignment_id).first()
        summary_rows = [
            ("Assignment", assignment.title),
            ("Deadline", str(assignment.deadline_at) if assignment.deadline_at else "No deadline"),
            ("Total Students", aa.total_targets if aa else 0),
            ("Submitted", aa.submitted_count if aa else 0),
            ("Missed", aa.missed_count if aa else 0),
            ("Late", aa.late_count if aa else 0),
            ("Completion Rate", f"{aa.completion_rate if aa else 0}%"),
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ]
        for r, (label, value) in enumerate(summary_rows, start=1):
            ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws2.cell(row=r, column=2, value=value)

        # save to temp file and upload to S3
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        bucket = os.getenv("S3_BUCKET_NAME")
        region = os.getenv("AWS_REGION")
        import uuid
        s3_key = f"exports/{job.requested_by}/{uuid.uuid4()}/tracker_{assignment.title[:30]}.xlsx"
        s3.upload_file(
            tmp_path, bucket, s3_key,
            ExtraArgs={"ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        )
        os.unlink(tmp_path)

        file_url = f"https://{bucket}.s3.{region}.amazonaws.com/{s3_key}"
        job.file_url = file_url
        job.status = "DONE"
        db.commit()

    except Exception as e:
        job.status = "FAILED"
        db.commit()
        raise
