from database import SessionLocal
from models.user import User
from models.class_ import Class, ClassMembership
from models.analytics import ClassAnalytics
from models.bulk_import import BulkImportBatch, BulkImportError
from utils.security import hash_password
from services.email_service import send_invite_email
from utils.id_generator import make_mentor_reg_id
import openpyxl
import os

def process_bulk_import(batch_id: str, file_path: str, admin_id: str):
    db = SessionLocal()
    b = db.query(BulkImportBatch).filter(BulkImportBatch.id == batch_id).first()
    if not b:
        db.close()
        return
        
    b.status = 'VALIDATING'
    db.commit()
    
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        b.status = 'FAILED'
        db.add(BulkImportError(batch_id=b.id, sheet_name='File', row_number=0, error_message=str(e)))
        db.commit()
        db.close()
        return
        
    class_map = {}
    tot = 0
    succ = 0
    fail = 0
    emails_to_send = []
    
    if "Classes" in wb.sheetnames:
        for i, row in enumerate(wb["Classes"].iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]: continue
            tot += 1
            try:
                cname = str(row[0])
                ext = db.query(Class).filter(Class.class_name == cname, Class.admin_id == admin_id).first()
                if ext:
                    class_map[cname] = str(ext.id)
                    succ += 1
                    continue
                c = Class(admin_id=admin_id, class_name=cname, description=row[1] if len(row)>1 else None, academic_year=str(row[2]) if len(row)>2 and row[2] else None, status='ACTIVE')
                db.add(c)
                db.flush()
                db.add(ClassAnalytics(class_id=c.id))
                class_map[cname] = str(c.id)
                succ += 1
            except Exception as e:
                fail += 1
                db.add(BulkImportError(batch_id=b.id, sheet_name='Classes', row_number=i, error_message=str(e)))
                
    if "Mentors" in wb.sheetnames:
        for i, row in enumerate(wb["Mentors"].iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]: continue
            tot += 1
            try:
                cname, name, email, pwd, is_prim = row[0], row[1], row[2], row[3], row[4]
                if cname not in class_map: raise ValueError(f"Class '{cname}' not found")
                if db.query(User).filter(User.email == email).first(): raise ValueError("Email exists")
                reg = make_mentor_reg_id()
                u = User(role='MENTOR', status='ACTIVE', full_name=name, email=email, password_hash=hash_password(str(pwd)), registration_id=reg)
                db.add(u)
                db.flush()
                db.add(ClassMembership(class_id=class_map[cname], user_id=u.id, member_role='MENTOR', is_primary_mentor=bool(is_prim), status='ACTIVE', joined_via='BULK_IMPORT'))
                emails_to_send.append((email, name, str(pwd), reg, cname))
                succ += 1
            except Exception as e:
                fail += 1
                db.add(BulkImportError(batch_id=b.id, sheet_name='Mentors', row_number=i, error_message=str(e)))
                
    if "Students" in wb.sheetnames:
        for i, row in enumerate(wb["Students"].iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]: continue
            tot += 1
            try:
                cname, name, email, pwd, reg = row[0], row[1], row[2], row[3], row[4]
                if cname not in class_map: raise ValueError(f"Class '{cname}' not found")
                if db.query(User).filter(User.email == email).first(): raise ValueError("Email exists")
                if db.query(User).filter(User.registration_id == str(reg)).first(): raise ValueError("Registration ID exists")
                u = User(role='STUDENT', status='ACTIVE', full_name=name, email=email, password_hash=hash_password(str(pwd)), registration_id=str(reg))
                db.add(u)
                db.flush()
                db.add(ClassMembership(class_id=class_map[cname], user_id=u.id, member_role='STUDENT', status='PENDING', joined_via='BULK_IMPORT'))
                emails_to_send.append((email, name, str(pwd), str(reg), cname))
                succ += 1
            except Exception as e:
                fail += 1
                db.add(BulkImportError(batch_id=b.id, sheet_name='Students', row_number=i, error_message=str(e)))
                
    b.total_rows = tot
    b.success_rows = succ
    b.failed_rows = fail
    b.status = 'COMPLETED' if fail == 0 else 'PARTIAL'
    db.commit()
    
    try: os.remove(file_path)
    except: pass
        
    for em in emails_to_send:
        try: send_invite_email(em[0], em[1], em[2], em[3], em[4])
        except: pass
            
    db.close()
