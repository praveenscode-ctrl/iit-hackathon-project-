import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\PRAVEEN\Desktop\AssignHub\backup_docs\template -2.xlsx")
print("Sheets:", wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    print(f"\nSheet {sheet} (total rows {len(rows)}):")
    for r in rows[:10]:
        print(r)
